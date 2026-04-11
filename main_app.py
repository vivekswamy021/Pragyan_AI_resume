# app.py

import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
import traceback
import re
from dotenv import load_dotenv 
from datetime import date 
import csv 
from streamlit.runtime.uploaded_file_manager import UploadedFile

# -------------------------
# CONFIGURATION & API SETUP
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"
DEFAULT_JOB_TYPES = ["Full-time", "Contract", "Internship", "Remote", "Part-time"]
DEFAULT_ROLES = ["Software Engineer", "Data Scientist", "Product Manager", "HR Manager", "Marketing Specialist", "Operations Analyst"]

load_dotenv()
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

if not GROQ_API_KEY:
    st.warning("🚨 WARNING: GROQ_API_KEY not set. AI functionality disabled.")
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set.")
            return Completions()
    client = MockGroqClient()
else:
    client = Groq(api_key=GROQ_API_KEY)

# -------------------------
# CORE UTILITY FUNCTIONS
# -------------------------

def go_to(page_name):
    st.session_state.page = page_name

def get_file_type(file_path):
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    return ext if ext in ['pdf', 'docx', 'xlsx', 'txt', 'json', 'csv'] else 'txt'

def extract_content(file_type, file_path):
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                text = '\n'.join([page.extract_text() or "" for page in pdf.pages])
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                for row in workbook[sheet].iter_rows(values_only=True):
                    text += ' | '.join([str(c) for c in row if c is not None]) + '\n'
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        return text if text.strip() else "Error: File is empty."
    except Exception as e:
        return f"Error: {e}"

# -------------------------
# AI LOGIC (REFINED JSON EXTRACTION)
# -------------------------

def parse_with_llm(text):
    if not GROQ_API_KEY: return {"error": "API Key missing"}
    prompt = f"""Extract details from this resume into JSON: Name, Email, Phone, Skills, 
    Education (list), Experience (list), summary (3 sentence bio), Github, LinkedIn.
    Resume: {text}"""
    try:
        response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.1)
        content = response.choices[0].message.content.strip()
        # Aggressive Regex Fix for "Extra Data" error
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            return json.loads(json_match.group(0))
        raise ValueError("No JSON found")
    except Exception as e:
        return {"error": str(e), "name": "Unknown"}

@st.cache_data(show_spinner="Matching...")
def evaluate_jd_fit(jd_text, parsed_json):
    if not GROQ_API_KEY: return "AI Disabled"
    prompt = f"Match this Resume JSON: {json.dumps(parsed_json)} against this JD: {jd_text}. Give Score/10 and Strengths/Gaps."
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}])
    return response.choices[0].message.content

def extract_jd_from_linkedin_url(url):
    return f"Simulated JD for role at {url}. Requirements: Python, SQL, Cloud."

# -------------------------
# ADMIN DASHBOARD
# -------------------------

def admin_dashboard():
    st.title("🧑‍💼 Admin Dashboard")
    if st.button("🚪 Log Out"): go_to("login"); st.rerun()

    tab1, tab2, tab3 = st.tabs(["📊 Resume Analysis", "🛠️ User Management", "📈 Statistics"])
    
    with tab1:
        files = st.file_uploader("Upload Resumes", accept_multiple_files=True)
        if st.button("Process Resumes"):
            for f in files:
                with tempfile.NamedTemporaryFile(delete=False, suffix=f".{get_file_type(f.name)}") as tmp:
                    tmp.write(f.getbuffer())
                    txt = extract_content(get_file_type(tmp.name), tmp.name)
                    res = parse_with_llm(txt)
                    st.session_state.resumes_to_analyze.append({"name": f.name, "parsed": res, "applied_jd": "Pending", "submitted_date": str(date.today())})
                    st.session_state.resume_statuses[f.name] = "Pending"
            st.success("Processed!")

    with tab2:
        st.subheader("Candidate Approval")
        for idx, r in enumerate(st.session_state.resumes_to_analyze):
            with st.container(border=True):
                st.write(f"**Candidate:** {r['name']} | **Status:** {st.session_state.resume_statuses[r['name']]}")
                st.write(f"Summary: {r['parsed'].get('summary', 'N/A')}")
                c1, c2, c3 = st.columns(3)
                if c1.button("✅ Approve", key=f"app_{idx}"): st.session_state.resume_statuses[r['name']] = "Approved"; st.rerun()
                if c2.button("❌ Reject", key=f"rej_{idx}"): st.session_state.resume_statuses[r['name']] = "Rejected"; st.rerun()
                if c3.button("🟡 Pending", key=f"pen_{idx}"): st.session_state.resume_statuses[r['name']] = "Pending"; st.rerun()

        st.divider()
        st.subheader("Vendor Approval")
        with st.form("vendor_form", clear_on_submit=True):
            v_name = st.text_input("Vendor Name")
            v_contact = st.text_input("Contact Person")
            if st.form_submit_button("Add Vendor"):
                st.session_state.vendors.append({"name": v_name, "contact": v_contact, "status": "Pending Review"})
                st.session_state.vendor_statuses[v_name] = "Pending Review"
                st.rerun()
        st.write(st.session_state.vendors)

    with tab3:
        st.write(f"Total Candidates: {len(st.session_state.resumes_to_analyze)}")
        st.write(f"Total Vendors: {len(st.session_state.vendors)}")

# -------------------------
# CANDIDATE DASHBOARD
# -------------------------

def candidate_dashboard():
    st.title("👩‍🎓 Candidate Dashboard")
    if st.button("🚪 Log Out"): go_to("login"); st.rerun()

    t1, t2 = st.tabs(["🚀 CV Management & Match", "📝 Education"])
    
    with t1:
        st.subheader("Match Your Resume")
        uploaded_file = st.file_uploader("Upload Resume")
        jd_text = st.text_area("Paste JD")
        if st.button("Run Analysis"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(uploaded_file.getbuffer())
                txt = extract_content("pdf", tmp.name)
                parsed = parse_with_llm(txt)
                fit = evaluate_jd_fit(jd_text, parsed)
                st.write(fit)

    with t2:
        st.subheader("Add Education")
        with st.form("edu_form", clear_on_submit=True):
            deg = st.selectbox("Degree", ["B.Tech", "M.Tech", "MBA", "B.Sc", "Other"])
            col = st.text_input("College")
            uni = st.text_input("University")
            dfrom = st.date_input("From", value=date(2020,1,1))
            dto = st.date_input("To", value=date.today())
            if st.form_submit_button("Add Education"):
                st.session_state.manual_education.append(f"{deg} at {col} ({uni}) [{dfrom} to {dto}]")
                st.rerun()
        st.write(st.session_state.manual_education)

# -------------------------
# MAIN ROUTING
# -------------------------

def main():
    st.set_page_config(layout="wide", page_title="PragyanAI Portal")

    # Initialize State
    if 'page' not in st.session_state: st.session_state.page = "login"
    if 'resumes_to_analyze' not in st.session_state: st.session_state.resumes_to_analyze = []
    if 'resume_statuses' not in st.session_state: st.session_state.resume_statuses = {}
    if 'vendors' not in st.session_state: st.session_state.vendors = []
    if 'vendor_statuses' not in st.session_state: st.session_state.vendor_statuses = {}
    if 'manual_education' not in st.session_state: st.session_state.manual_education = []

    if st.session_state.page == "login":
        st.title("PragyanAI Job Portal")
        role = st.selectbox("I am a...", ["Admin", "Candidate", "Hiring Company"])
        if st.button("Login"):
            if role == "Admin": go_to("admin")
            elif role == "Candidate": go_to("candidate")
            else: go_to("hiring")
            st.rerun()
    
    elif st.session_state.page == "admin": admin_dashboard()
    elif st.session_state.page == "candidate": candidate_dashboard()
    elif st.session_state.page == "hiring": st.title("🏢 Hiring Dashboard (Coming Soon)"); st.button("Back", on_click=lambda: go_to("login"))

if __name__ == "__main__":
    main()
