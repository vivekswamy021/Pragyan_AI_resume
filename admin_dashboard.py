import streamlit as st
import os
import pdfplumber
import docx
import openpyxl
import json
import tempfile
from groq import Groq
import traceback
import re 
from dotenv import load_dotenv 
from datetime import date 
import csv 
from streamlit.runtime.uploaded_file_manager import UploadedFile

# -------------------------
# CONFIGURATION & API SETUP
# -------------------------

GROQ_MODEL = "llama-3.1-8b-instant"

load_dotenv()
GROQ_API_KEY = os.getenv('GROQ_API_KEY')

if not GROQ_API_KEY:
    class MockGroqClient:
        def chat(self):
            class Completions:
                def create(self, **kwargs):
                    raise ValueError("GROQ_API_KEY not set. AI functions disabled.")
            return Completions()
    client = MockGroqClient()
else:
    client = Groq(api_key=GROQ_API_KEY)

# -------------------------
# UTILITY & EXTRACTION FUNCTIONS
# -------------------------

def go_to(page_name):
    st.session_state.page = page_name

def get_file_type(file_path):
    ext = os.path.splitext(file_path)[1].lower().strip('.')
    if ext == 'pdf': return 'pdf'
    elif ext == 'docx': return 'docx'
    elif ext == 'xlsx': return 'xlsx'
    else: return 'txt' 

def extract_content(file_type, file_path):
    text = ''
    try:
        if file_type == 'pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text: text += page_text + '\n'
        elif file_type == 'docx':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
        elif file_type == 'xlsx':
            workbook = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(c) for c in row if c is not None])
                    if row_text.strip(): text += row_text + '\n'
        else:
             with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        return text if text.strip() else f"Error: {file_type.upper()} extraction failed."
    except Exception as e:
        return f"Fatal Extraction Error: {e}"

@st.cache_data(show_spinner="Extracting JD metadata...")
def extract_jd_metadata(jd_text):
    if not GROQ_API_KEY: return {"role": "N/A", "job_type": "N/A", "key_skills": []}
    prompt = f"Analyze this JD and extract role, job_type, and key_skills in JSON: {jd_text}"
    try:
        response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.0)
        json_match = re.search(r'\{.*\}', response.choices[0].message.content.strip(), re.DOTALL)
        parsed = json.loads(json_match.group(0))
        return {
            "role": parsed.get("role", "General Analyst"),
            "job_type": parsed.get("job_type", "Full-time"),
            "key_skills": parsed.get("key_skills", [])
        }
    except: return {"role": "General Analyst", "job_type": "Full-time", "key_skills": []}

@st.cache_data(show_spinner="Analyzing content...")
def parse_with_llm(text, return_type='json'):
    if not GROQ_API_KEY: return {"error": "API key missing."}
    prompt = f"Extract Name, Email, Phone, Skills, Education, Experience, Projects from this resume. Also include a 'summary' key (3 sentences). Text: {text}"
    try:
        response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.2)
        json_match = re.search(r'\{.*\}', response.choices[0].message.content.strip(), re.DOTALL)
        return json.loads(json_match.group(0))
    except Exception as e: return {"error": f"LLM error: {e}"}

def evaluate_jd_fit(job_description, parsed_json):
    if not GROQ_API_KEY or "error" in parsed_json: return "Analysis failed."
    prompt = f"Evaluate fit score /10 and match % for Skills, Experience, Education. JD: {job_description} Resume: {json.dumps(parsed_json)}"
    response = client.chat.completions.create(model=GROQ_MODEL, messages=[{"role": "user", "content": prompt}], temperature=0.3)
    return response.choices[0].message.content.strip()

def parse_and_store_resume(file_input, file_name_key='default', source_type='file'):
    if source_type == 'file':
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file_input.name) 
        with open(temp_path, "wb") as f: f.write(file_input.getbuffer()) 
        file_type = get_file_type(temp_path)
        text = extract_content(file_type, temp_path)
        file_name = file_input.name
    else:
        text = file_input
        file_name = "Pasted Resume"
    parsed = parse_with_llm(text)
    return {"parsed": parsed, "full_text": text, "name": parsed.get('name', file_name)}

def update_resume_metadata(resume_name, new_status, applied_jd, submitted_date, resume_list_index):
    st.session_state.resume_statuses[resume_name] = new_status
    if 0 <= resume_list_index < len(st.session_state.resumes_to_analyze):
        st.session_state.resumes_to_analyze[resume_list_index]['applied_jd'] = applied_jd
        st.session_state.resumes_to_analyze[resume_list_index]['submitted_date'] = submitted_date
        st.toast(f"Updated **{resume_name}** to **{new_status}**.")

# -------------------------
# DASHBOARD TABS
# -------------------------

def candidate_approval_tab_content():
    st.header("👤 Candidate Approval")
    if not st.session_state.resumes_to_analyze:
        st.info("No resumes uploaded yet.")
        return
    
    jd_options = [item['name'] for item in st.session_state.admin_jd_list]
    jd_options.insert(0, "Select JD") 

    for idx, resume_data in enumerate(st.session_state.resumes_to_analyze):
        resume_name = resume_data['name']
        current_status = st.session_state.resume_statuses.get(resume_name, "Pending")
        parsed = resume_data.get('parsed', {})
        
        with st.container(border=True):
            st.markdown(f"### **Candidate:** {resume_name} (Status: **{current_status}**)")
            c1, c2 = st.columns(2)
            c1.markdown(f"**📧 Email:** `{parsed.get('email', 'N/A')}`\n**📱 Phone:** `{parsed.get('phone', 'N/A')}`")
            c2.markdown(f"**🎓 Education:** `{parsed.get('education', ['N/A'])[0] if isinstance(parsed.get('education'), list) else 'N/A'}`")
            st.markdown(f"**Brief Info:** *{parsed.get('summary', 'No summary available.')}*")
            
            sc1, sc2 = st.columns(2)
            applied_jd = sc1.selectbox("Applied for JD", jd_options, key=f"jd_{idx}")
            sub_date = sc2.date_input("Submission Date", value=date.today(), key=f"date_{idx}")
            
            b1, b2, b3, _ = st.columns([1, 1, 1, 5])
            jd_final = applied_jd if applied_jd != "Select JD" else "N/A"
            dt_final = sub_date.strftime("%Y-%m-%d")

            if b1.button("✅ Approve", key=f"app_{idx}"):
                update_resume_metadata(resume_name, "Approved", jd_final, dt_final, idx)
                st.rerun()
            if b2.button("❌ Reject", key=f"rej_{idx}"):
                update_resume_metadata(resume_name, "Rejected", jd_final, dt_final, idx)
                st.rerun()
            if b3.button("🟡 Pending", key=f"pen_{idx}"):
                update_resume_metadata(resume_name, "Pending", jd_final, dt_final, idx)
                st.rerun()

def vendor_approval_tab_content():
    st.header("🤝 Vendor Approval") 
    with st.form("add_vendor", clear_on_submit=True):
        st.markdown("#### Vendor Details")
        v_name = st.text_input("Vendor Company Name")
        v_person = st.text_input("Contact Person")
        v_email = st.text_input("Email ID")
        v_submitted = st.form_submit_button("Add Vendor")
        if v_submitted and v_name and v_person and v_email:
            st.session_state.vendors.append({'name': v_name, 'person': v_person, 'email': v_email, 'date': date.today().strftime("%Y-%m-%d")})
            st.session_state.vendor_statuses[v_name] = "Pending Review"
            st.success(f"Added {v_name}.")
            st.rerun()

    st.markdown("---")
    st.subheader("Update Vendor Status")
    for idx, vendor in enumerate(st.session_state.vendors):
        name = vendor['name']
        with st.container(border=True):
            st.write(f"**{name}** | Contact: {vendor['person']} ({vendor['email']})")
            new_v_stat = st.selectbox("Status", ["Pending Review", "Approved", "Rejected"], key=f"vstat_{idx}", index=["Pending Review", "Approved", "Rejected"].index(st.session_state.vendor_statuses[name]))
            if st.button("Update Status", key=f"vup_{idx}"):
                st.session_state.vendor_statuses[name] = new_v_stat
                st.rerun()

def admin_dashboard():
    st.title("🧑‍💼 Admin Dashboard")
    if st.button("🚪 Log Out"): go_to("login")
    
    # Initialize session states
    if "admin_jd_list" not in st.session_state: st.session_state.admin_jd_list = []
    if "resumes_to_analyze" not in st.session_state: st.session_state.resumes_to_analyze = []
    if "resume_statuses" not in st.session_state: st.session_state.resume_statuses = {}
    if "vendors" not in st.session_state: st.session_state.vendors = []
    if "vendor_statuses" not in st.session_state: st.session_state.vendor_statuses = {}

    tab_jd, tab_analysis, tab_user, tab_stats = st.tabs(["📄 JD Management", "📊 Resume Analysis", "🛠️ User Management", "📈 Statistics"])

    with tab_jd:
        st.subheader("Manage Job Descriptions")
        jd_txt = st.text_area("Paste JD Text Here")
        if st.button("Add JD"):
            meta = extract_jd_metadata(jd_txt)
            st.session_state.admin_jd_list.append({"name": f"JD {len(st.session_state.admin_jd_list)+1}", "content": jd_txt, **meta})
            st.success("JD Added.")

    with tab_analysis:
        st.subheader("Resume Batch Analysis")
        files = st.file_uploader("Upload Resumes", accept_multiple_files=True)
        if st.button("Parse Resumes"):
            for f in files:
                res = parse_and_store_resume(f)
                st.session_state.resumes_to_analyze.append(res)
                st.session_state.resume_statuses[res['name']] = "Pending"
            st.success("Resumes Parsed.")

    with tab_user:
        ut1, ut2 = st.tabs(["Candidates", "Vendors"])
        with ut1: candidate_approval_tab_content()
        with ut2: vendor_approval_tab_content()

    with tab_stats:
        st.header("📊 System Statistics")
        st.markdown("---")
        
        # Candidate Stats
        st.subheader("Candidate Status Breakdown")
        c_stats = {"Pending": 0, "Approved": 0, "Rejected": 0}
        for s in st.session_state.resume_statuses.values(): c_stats[s] = c_stats.get(s, 0) + 1
        col1, col2, col3 = st.columns(3)
        col1.metric("Pending Candidates", c_stats["Pending"])
        col2.metric("Approved Candidates", c_stats["Approved"])
        col3.metric("Rejected Candidates", c_stats["Rejected"])
        
        st.markdown("---")
        
        # Vendor Stats
        st.subheader("Vendor Status Breakdown")
        v_stats = {"Pending Review": 0, "Approved": 0, "Rejected": 0}
        for s in st.session_state.vendor_statuses.values(): v_stats[s] = v_stats.get(s, 0) + 1
        vcol1, vcol2, vcol3 = st.columns(3)
        vcol1.metric("Pending Vendors", v_stats["Pending Review"])
        vcol2.metric("Approved Vendors", v_stats["Approved"])
        vcol3.metric("Rejected Vendors", v_stats["Rejected"])

# Run logic
if __name__ == "__main__":
    if "page" not in st.session_state: st.session_state.page = "admin_dashboard"
    if st.session_state.page == "admin_dashboard":
        admin_dashboard()
